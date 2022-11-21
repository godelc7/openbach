#!/usr/bin/env python3

# OpenBACH is a generic testbed able to control/configure multiple
# network/physical entities (under test) and collect data from them. It is
# composed of an Auditorium (HMIs), a Controller, a Collector and multiple
# Agents (one for each network entity that wants to be tested).
#
#
# Copyright © 2016-2020 CNES
#
#
# This file is part of the OpenBACH testbed.
#
#
# OpenBACH is a free software : you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, either version 3 of the License, or (at your option) any later
# version.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY, without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program. If not, see http://www.gnu.org/licenses/.


"""Provide time period summary of data generated by OpenBACH jobs"""


__author__ = 'Viveris Technologies'
__credits__ = '''Contributors:
 * Aichatou Garba Abdou <aichatou.garba-abdou@viveris.fr>
'''


import itertools
import tempfile
import argparse
import syslog
import os

import pandas as pd
from datetime import datetime
from openpyxl import Workbook,load_workbook
from dateutil.parser import parse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Side, Border,Font,PatternFill



import collect_agent
from data_access.post_processing import Statistics

UNIT_OPTION={'s', 'ms' ,'bits/s', 'Kbits/s', 'Mbits/s','Gbits/s','Bytes' ,'KBytes', 'MBytes', 'GBytes'}


def worksheet_style(worksheet,title):

   worksheet.title=title
   fill_color=PatternFill(fill_type='solid',start_color='00333399',end_color='FF000000')
   border_color =Side(border_style='thin', color='000000')  
   for index,column in enumerate(worksheet.columns,1):
      worksheet['{}1'.format(get_column_letter(index))].fill=fill_color
      worksheet['{}1'.format(get_column_letter(index))].font=Font(size=12,bold=True,color='00FFFFFF')
      for cell in column:
         cell.border = Border(top=border_color, bottom=border_color, left=border_color, right=border_color)
         cell.alignment=Alignment(horizontal='center',vertical='center')
      worksheet.row_dimensions[index].height = 30
      worksheet.column_dimensions[get_column_letter(index)].width = 20
   for index,_ in enumerate(worksheet.rows,1):
       worksheet['{}{}'.format(get_column_letter(1),index)].fill=fill_color
       worksheet['{}{}'.format(get_column_letter(1),index)].font=Font(size=12,bold=True,color='00FFFFFF')

def get_evol_value(path_to_file,field,stats):

   workbook_ref=load_workbook(path_to_file)
   sheet_ref=workbook_ref[field]
   stats_ref=[]
   col_letter=_get_column_letter(sheet_ref,stats)

 
   for cell in sheet_ref[col_letter]:
        if cell.value != stats:
                stats_ref.append(cell.value)
   
   return stats_ref

def _get_column_letter(worksheet,value):
   for num,col in enumerate(worksheet.columns,1):
      for cell in col:
          if cell.value==value:
            column_letter=get_column_letter(num)
   return column_letter

def reference_style(worksheet,percent,column_title):

   column_letter=_get_column_letter(worksheet,column_title)
   for index,cell in enumerate(worksheet[column_letter],1):
      if cell.value !=column_title:
         if percent <= 33:
            worksheet['{}{}'.format(column_letter,index)].font=Font(color="ff0000")
         elif percent>34 and percent <= 66:
            worksheet['{}{}'.format(column_letter,index)].font=Font(color="ff7f00")
         elif percent > 67:     
            worksheet['{}{}'.format(column_letter,index)].font=Font(color="00ff00")

def get_evol_state(level,value,value_ref):
   palier=(value_ref*level)/100
   diff=value-value_ref
   evol=(diff*100)/value_ref
   if abs(diff)<= palier:
        state= '\u268C' #Stable
   if diff > palier:
        state= '\u2197' #En hausse
   if diff < (-palier):
        state='\u2198'  #En baisse
   return evol,state

def multiplier(base, unit):

        if unit == base:
                return 1
        if unit.startswith('GBytes'):
                return 1024 * 1024 * 1024
        if unit.startswith('MBytes'):
                return 1024 * 1024
        if unit.startswith('KBytes'):
                return 1024
        if unit.startswith('m'):
                return 0.001
        if unit.startswith('s'):
                return 1000
        if unit.startswith('Gbits'):
                return 1000 * 1000 * 1000
        if unit.startswith('Mbits'):
                return 1000 * 1000
        if unit.startswith('Kbits'):
                return 1000

        return 1

def main(
        agent_name,job_name, statistic_name,begin_date,end_date,start_journey,start_evening,start_night,reference,
        level,path_to_file,stat_unit,table_unit,stat_title,compute_median,compute_mean, stats_with_suffixes):

    file_ext='xlsx'
    statistics = Statistics.from_default_collector()
    statistics.origin = 0
    with tempfile.TemporaryDirectory(prefix='openbach-summary_time_period-') as root:       
        

        if not begin_date and not end_date:
                timestamp=None
        else:
                begin_date=parse(begin_date)
                end_date=parse(end_date)
                timestamp=[int(datetime.timestamp(begin_date)*1000),int(datetime.timestamp(end_date)*1000)]

        
        data_collection = statistics.fetch_all(
                job=job_name,agent=agent_name,
                suffix = None if stats_with_suffixes else '',
                fields=[statistic_name],timestamps=timestamp)
        
        df=data_collection.dataframe

        workbook=Workbook() 
        worksheet=workbook.active 
        
        if not level:
                level=10
        if not start_journey :
                start_journey=7
        if not start_evening:
                start_evening=18
        if not start_night :
                start_night=0 

        df.index=pd.to_datetime(df.index,unit='ms')

        if not table_unit:
                table_unit=''
        facteur=multiplier(stat_unit,table_unit)

        means,_=data_collection.compute_function(df,'moyenne',facteur,start_journey,start_evening,start_night)
        medians,_=data_collection.compute_function(df,'mediane',facteur,start_journey,start_evening,start_night)
        moments=list(means.index)
        means=means.round(2)
        medians=medians.round(2)
        if not path_to_file :

                if not compute_mean and compute_median:
                
                   worksheet.append([f'{statistic_name} {table_unit}','Médiane','% Médiane Cible'])
                   for (moment ,median) in  zip(moments,list(medians)):
                        percent=(median*100)/reference
                        worksheet.append([moment,median,2,f'{percent}%'])
                        reference_style(worksheet,percent,'% Médiane Cible')

                if not compute_median and compute_mean:
                   worksheet.append([f'{statistic_name} {table_unit}','Moyenne','% Moyenne Cible']) 
                   for (moment, mean ) in  zip(moments,list(means)):
                        percent=(mean*100)/reference
                        worksheet.append([moment,mean,f'{percent}%'])
                        reference_style(worksheet,percent,'% Moyenne Cible')

                if compute_mean and compute_median:
                
                   worksheet.append([f'{statistic_name} {table_unit}','Moyenne','% Moyenne Cible','Médiane','% Médiane Cible'])
                
                   for (moment, mean ,median) in  zip(moments,list(means),list(medians)):
                        percent_median=(median*100)/reference
                        percent_mean=(mean*100)/reference
                        worksheet.append([moment,mean,f'{percent_mean}%',median,f'{percent_median}%'])
                        reference_style(worksheet,percent_mean,'% Moyenne Cible')
                        reference_style(worksheet,percent_median,'% Médiane Cible')

        else:

                if not compute_mean and compute_median:

                        medians_ref=get_evol_value(path_to_file,statistic_name,'Médiane')
                        worksheet.append([f'{statistic_name} {table_unit}','Médiane','% Médiane Cible','Evolution de la Médiane'])
                        for (moment ,median,median_ref) in  itertools.zip_longest(moments,list(medians),medians_ref,fillvalue=0):
                           percent=(median*100)/reference
                           if median_ref!=0:
                                evol,state=get_evol_state(level,median,median_ref)
                                row=[moment,median,f'{percent}%', f'{state} {evol}%']
                           else:
                                row=[moment,median,f'{percent}%','NaN']
                        worksheet.append(row)
                        reference_style(worksheet,percent,'% Médiane Cible')

                if not compute_median and compute_mean:

                        means_ref=get_evol_value(path_to_file,statistic_name,'Moyenne')                          
                        worksheet.append([f'{statistic_name} {table_unit}','Moyenne','% Moyenne Cible','Evolution de Moyenne']) 
                        for (moment, mean,mean_ref ) in  itertools.zip_longest(moments,list(means),means_ref,fillvalue=0):
                           percent=(mean*100)/reference
                           if mean_ref!=0:
                                evol,state=get_evol_state(level,mean,mean_ref)
                                row=[moment,mean,f'{percent}%', f'{state} {evol}%']
                           else:
                                row=[moment,mean,f'{percent}%','NaN']
                        worksheet.append(row)
                        reference_style(worksheet,percent,'% Moyenne Cible')

                if compute_mean and compute_median:

                        medians_ref=get_evol_value(path_to_file,statistic_name,'Médiane') 
                        means_ref=get_evol_value(path_to_file,statistic_name,'Moyenne')                               
                        worksheet.append([f'{statistic_name} {table_unit}','Moyenne','% Moyenne Cible','Evolution de Moyenne','Médiane','% Médiane Cible','Evolution de Médiane'])
                        for (moment, mean ,median,mean_ref,median_ref) in   itertools.zip_longest(moments,list(means),list(medians),means_ref,medians_ref,fillvalue=0):
                           percent_median=(median*100)/reference
                           percent_mean=(mean*100)/reference
                           if mean_ref!=0 and median_ref!=0:              
                                evol_mean,state_mean=get_evol_state(level,mean,mean_ref)
                                evol_median,state_median=get_evol_state(level,median,median_ref) 
                                row=[moment,mean,f'{percent_mean}%', f'{state_mean} {evol_mean}%',
                                        median,f'{percent_median}%', f'{state_median} {evol_median}%']
                           else:
                                row=[moment,mean,f'{percent_mean}%','NaN',median,f'{percent_median}%','NaN']
                        worksheet.append(row)
                        reference_style(worksheet,percent_mean,'% Moyenne Cible')
                        reference_style(worksheet,percent_median,'% Médiane Cible')
                
        if not stat_title:
                stat_title =statistic_name
        worksheet_style(worksheet,stat_title)     
                
        filepath = os.path.join(root, 'summary_time_period_{}.{}'.format(statistic_name, file_ext))    
        workbook.save(filepath)
        collect_agent.store_files(collect_agent.now(), figure=filepath)

if __name__ == '__main__':
    with collect_agent.use_configuration('/opt/openbach/agent/jobs/summary_time_period/summary_time_period_rstats_filter.conf'):
        parser = argparse.ArgumentParser(description=__doc__)
        parser.add_argument(
                metavar='AGENT_NAME',dest='agents',
                type=str, default=[],
                help='Agent name to fetch data from')
        parser.add_argument(
                metavar='JOB_NAME', type=str, default=[],dest='jobs',
                help='job name to fetch data from')
        parser.add_argument(
                metavar='STATISTIC' ,dest='statistics',
                default=[],help='statistics names to be analysed')
        parser.add_argument(
                metavar='REFERENCE',dest='reference',type=int,
                default=[],help='Reference value for comparison in desired stat unit')
        parser.add_argument(
                '-b', '--begin-date',metavar='BIGIN_DATE' ,dest='begin_date',
                default=[],help='Start date in format YYYY:MM:DD hh:mm:ss')
        parser.add_argument(
                '-e', '--end-date',metavar='END_DATE',dest='end_date',
                default=[],help='End date in format YYYY:MM:DD hh:mm:ss')

        parser.add_argument(
                '-sj', '--start-journey',metavar='START_JOURNEY' ,dest='start_journey',
                default=[],help='starting time of the day')
        parser.add_argument(
                '-se', '--start-evening',metavar='START_EVENING' ,dest='start_evening',
                default=[],help='starting time of the evening')
        parser.add_argument(
                '-sn', '--start-night',metavar='START_NIGHT',dest='start_night',
                default=[],help='starting time of the night')
        parser.add_argument(
                '-l', '--level',metavar='LEVEL',dest='level',type=int,
                default=[],help='Percentage level to observe the evolution')     
        parser.add_argument(
                '-p', '--path_to_file',metavar='PATH_TO_FILE',dest='path_to_file',
                default=[],help='Path to xlsx file for evolution calculation')
        parser.add_argument(
                '-ub', '--stat-unit', dest='stat_units',choices=UNIT_OPTION,
                metavar='STAT_UNIT',  default=[],
                help='Unit of the statistic')
        parser.add_argument(
                '-u', '--table-unit', dest='table_units',choices=UNIT_OPTION,
                metavar='TALE_UNIT', default=[],
                help='Unit to show on the table')
        parser.add_argument(
                '-t', '--stat-title', '--stat-title', dest='stat_title',
                metavar='STAT-TITLE', default=[],
                help='statistics names to display on the table')
        parser.add_argument(
                '-w', '--no-suffix', action='store_true',
                help='Do not plot statistics with suffixes')
        parser.add_argument(
                '--no_median',help='Do not compute median', action='store_true')
        parser.add_argument(
                '--no_mean',help='Do not compute mean', action='store_true')

        args = parser.parse_args()
        compute_median = not args.no_median
        compute_mean = not args.no_mean
        stats_with_suffixes = not args.no_suffix
        
        main(
            args.agents,args.jobs, args.statistics,args.begin_date,args.end_date,args.start_journey,args.start_evening,args.start_night,args.reference,
            args.level,args.path_to_file,args.stat_units,args.table_units,args.stat_title ,compute_median,compute_mean,stats_with_suffixes)
